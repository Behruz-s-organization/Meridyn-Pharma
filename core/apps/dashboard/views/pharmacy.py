import json

# django
from django.db.models import Q
from django.http import HttpResponse

# rest framework
from rest_framework import viewsets, permissions, views
from rest_framework.decorators import action

# drf yasg
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema

# openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# dashboard
from core.apps.dashboard.serializers import pharmacy as serializers

# shared
from core.apps.shared.models import Pharmacy
from core.apps.shared.utils.response_mixin import ResponseMixin


class PharmacyViewSet(viewsets.GenericViewSet, ResponseMixin):
    permission_classes = [permissions.IsAdminUser]
    queryset = Pharmacy.objects.all()

    def get_serializer_class(self):
        if self.action == "post": 
            return serializers.AdminPharmacyCreateSerializer
        elif self.action in ["patch", "put"]:
            return serializers.PharmacyUpdateSerializer
        else:
            return serializers.PharmacyListSerializer

    @swagger_auto_schema(
        tags=['Admin Pharmacies'],
        manual_parameters=[
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="name",
                description="name bo'yicha search",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="place",
                description="obyekt name bo'yicha search",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="district",
                description="tuman name bo'yicha search",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="user",
                description="qo'shgan foydalanuvchini ism va familiyasi bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                name='user_id',
                description="user id bo'yicha filter",
                required=False,
            ) 
        ],
    )
    @action(detail=False, methods=['get'], url_path="list")
    def get(self, request):
        try:
            # params
            name = request.query_params.get('name', None)
            place_name = request.query_params.get('place', None)
            district_name = request.query_params.get('district', None)
            user_full_name = request.query_params.get('user', None)
            user_id = request.query_params.get('user_id', None)

            queryset = self.queryset.all()

            # filters
            if name is not None:
                queryset = queryset.filter(name__istartswith=name)
            
            if district_name is not None:
                queryset = queryset.filter(district__name__istartswith=district_name)
            
            if place_name is not None:
                queryset = queryset.filter(place__name__istartswith=place_name)

            if user_full_name is not None:
                queryset = queryset.filter(
                    Q(user__first_name__istartswith=user_full_name) |
                    Q(user__last_name__istartswith=user_full_name) 
                )
            if not user_id is None:
                queryset = queryset.filter(user__id=user_id)


            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.success_response(
                    data=self.get_paginated_response(serializer.data).data,
                    message='malumotlar fetch qilindi'
                )
            serializer = self.get_serializer(queryset, many=True)
            return self.success_response(
                data=serializer.data,
                message='malumotlar fetch qilindi'
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )

    @swagger_auto_schema(
        tags=['Admin Pharmacies']
    )
    @action(detail=False, methods=['post'], url_path='create')
    def post(self, request):
        try:
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                obj = serializer.save()
                return self.success_response(
                    data=serializers.PharmacyListSerializer(obj).data,
                    message='malumot qoshildi'
                )
            return self.failure_response(
                data=serializer.errors,
                message='malumot qoshilmadi'
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )
    
    @swagger_auto_schema(
        tags=['Admin Pharmacies']
    )
    @action(detail=True, methods=['patch'], url_path='update')
    def update_pharmacy(self, request, pk=None):
        try:
            pharmacy = Pharmacy.objects.filter(id=pk).first()
            if not pharmacy:
                return self.failure_response(
                    data={},
                    message="pharmacy topilmadi",
                    status_code=404
                )
            serializer = self.get_serializer(data=request.data, instance=pharmacy)
            if serializer.is_valid():
                obj = serializer.save()
                return self.success_response(
                    data=serializers.PharmacyListSerializer(obj).data,
                    message='malumot tahrirlandi'
                )
            return self.failure_response(
                data=serializer.errors,
                message='malumot tahrirlandi'
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )
    
    @swagger_auto_schema(
        tags=['Admin Pharmacies']
    )
    @action(detail=True, methods=['delete'], url_path='delete')
    def delete(self, request, pk=None):
        try:
            pharmacy = Pharmacy.objects.filter(id=pk).first()
            if not pharmacy:
                return self.failure_response(
                    data={},
                    message="pharmacy topilmadi",
                    status_code=404
                )
            pharmacy.delete()
            return self.success_response(
                data={},
                message='malumot ochirildi',
                status_code=204
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )
    
    

class PharmacyExportView(views.APIView, ResponseMixin):
    permission_classes = [permissions.IsAuthenticated]

    @swagger_auto_schema(
        tags=['Admin Pharmacies']
    )    
    def get(self, request):
        try:
            queryset = Pharmacy.objects.select_related(
                'district', 'place', 'user'
            ).all()
            queryset = self.filter_queryset(request, queryset)
            response = self.create_excel_file(queryset)
            return response
        except Exception as e:
            return self.error_response(
                data=str(e),
            )
    
    def filter_queryset(self, request, queryset):
        district_id = request.query_params.get('district_id')
        if district_id:
            queryset = queryset.filter(district_id=district_id)
        
        place_id = request.query_params.get('place_id')
        if place_id:
            queryset = queryset.filter(place_id=place_id)
        
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(inn__icontains=search)
            )
        
        user_id = request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        return queryset
    
    def create_excel_file(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dorixonalar"
        
        header_fill = PatternFill(
            start_color="366092", 
            end_color="366092", 
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = [
            'ID',
            'Nomi',
            'INN',
            'Egasining telefoni',
            "Mas'ul shaxs telefoni",
            'Tuman',
            'Joy',
            'Foydalanuvchi',
            'Longitude',
            'Latitude',
            "Qo'shimcha manzil"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row_num, pharmacy in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1).value = pharmacy.id
            ws.cell(row=row_num, column=2).value = pharmacy.name
            ws.cell(row=row_num, column=3).value = pharmacy.inn
            ws.cell(row=row_num, column=4).value = pharmacy.owner_phone or ''
            ws.cell(row=row_num, column=5).value = pharmacy.responsible_phone or ''
            ws.cell(row=row_num, column=6).value = pharmacy.district.name if pharmacy.district else ''
            ws.cell(row=row_num, column=7).value = pharmacy.place.name if pharmacy.place else ''
            ws.cell(row=row_num, column=8).value = pharmacy.user.username if pharmacy.user else ''
            ws.cell(row=row_num, column=9).value = pharmacy.longitude
            ws.cell(row=row_num, column=10).value = pharmacy.latitude
            
            extra_loc = pharmacy.extra_location
            if extra_loc:
                ws.cell(row=row_num, column=11).value = json.dumps(
                    extra_loc, 
                    ensure_ascii=False
                )
            else:
                ws.cell(row=row_num, column=11).value = ''
        
        column_widths = [8, 30, 15, 18, 18, 20, 20, 20, 12, 12, 30]
        for i, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[column_letter].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=dorixonalar.xlsx'
        
        wb.save(response)
        
        return response