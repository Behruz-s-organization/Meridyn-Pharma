import json

# django
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.http import HttpResponse

# rest framework
from rest_framework import viewsets, permissions, views
from rest_framework.decorators import action

# drf yasg
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema

# openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# dashboard
from core.apps.dashboard.serializers import doctor as serializers

# shared
from core.apps.shared.models import Doctor
from core.apps.shared.utils.response_mixin import ResponseMixin


class DoctorViewSet(viewsets.GenericViewSet, ResponseMixin):
    permission_classes = [permissions.IsAdminUser]
    queryset = Doctor.objects.all()

    def get_serializer_class(self):
        if self.action == "post": 
            return serializers.DoctorCreateSerializer
        elif self.action in ("patch", "put"):
            return serializers.DoctorUpdateSerializer
        else:
            return serializers.DoctorListSerializer

    @swagger_auto_schema(
        tags=['Admin Doctors'],
        manual_parameters=[
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="full_name",
                description="shiforkor ism va familiyasi bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="district_name",
                description="tuman nomi bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="place_name",
                description="obyekt bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="work_place",
                description="ish joyi bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="sphere",
                description="lavozimi bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                name="user",
                description="qo'shgan foydalanuvchini ism va familiyasi bo'yicha qidirish",
                required=False,
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                name='user_id',
                description="user id bo'yicha filter",
                required=False,
            ) 
        ],
        operation_description="Shifokorlar ro'yxatini olish",
        operation_summary="Shifokolar ro'yxati",
        responses={
            200: openapi.Response(
                schema=None,
                description="Success",
                examples={
                    "application/json": {
                        "status_code": 200,
                        "status": "success",
                        "message": "malumotlar fetch qilindi",
                        "data": {
                            "count": 0,
                            "next": "string",
                            "pervious": "string",
                            "response": [
                                {
                                    "id": 0,
                                    "first_name": "string",
                                    "last_name": "string",
                                    "phone_number": "string",
                                    "work_place": "string",
                                    "sphere": "string",
                                    "description": "string",
                                    "district": {
                                        "id": 0,
                                        "name": "string",
                                    },
                                    "place": {
                                        "id": 0,
                                        "name": "string"
                                    },
                                    "user": {
                                        'id': 0,
                                        'first_name': "string",
                                        'last_name': "string",
                                    },
                                    "longitude": 0.0,
                                    "latitude": 0.0,
                                    "extra_location": {},
                                    "created_at": "string"
                                }
                            ]
                        }
                    }
                }
            ),
            500: openapi.Response(
                schema=None,
                description="Server Error",
                examples={
                    "application/json": {
                        "status_code": 500,
                        "success": "error",
                        "message": "xatolik",
                        "data": "string"
                    }
                }
            )
        }
    )
    @action(detail=False, methods=['get'], url_path="list")
    def get(self, request):
        try:
            # params
            full_name = request.query_params.get('full_name', None)
            district_name = request.query_params.get('district_name', None)
            place_name = request.query_params.get('place_name', None)
            work_place = request.query_params.get('work_place', None)
            sphere = request.query_params.get('sphere', None)
            user_full_name = request.query_params.get('user', None)
            user_id = request.query_params.get('user_id', None)


            queryset = self.queryset.all()

            # filters
            if full_name is not None:
                queryset = queryset.filter(
                    Q(first_name__istartswith=full_name) |
                    Q(last_name__istartswith=full_name) 
                )
            if district_name is not None:
                queryset = queryset.filter(
                    district__name__istartswith=district_name
                )
            if place_name is not None:
                queryset = queryset.filter(
                    place__name__istartswith=place_name
                )
            if work_place is not None:
                queryset = queryset.filter(
                    work_place__istartswith=work_place
                )
            if sphere is not None:
                queryset = queryset.filter(
                    sphere__istartswith=sphere
                )
            if user_full_name is not None:
                queryset = queryset.filter(
                    Q(user__first_name__istartswith=user_full_name) |
                    Q(user__last_name__istartswith=user_full_name) 
                )
            if not user_id is None:
                queryset = queryset.filter(user__id=user_id)

            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.success_response(
                    data=self.get_paginated_response(serializer.data).data,
                    message='malumotlar fetch qilindi'
                )
            serializer = self.get_serializer(queryset, many=True)
            return self.success_response(
                data=serializer.data,
                message='malumotlar fetch qilindi'
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )

    @swagger_auto_schema(
        tags=['Admin Doctors']
    )
    @action(detail=False, methods=['post'], url_path='create')
    def post(self, request):
        try:
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                obj = serializer.save()
                return self.success_response(
                    data=serializers.DoctorListSerializer(obj).data,
                    message='malumot qoshildi'
                )
            return self.failure_response(
                data=serializer.errors,
                message='malumot qoshilmadi'
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )
    
    @swagger_auto_schema(
        tags=['Admin Doctors']
    )
    @action(detail=True, methods=['patch'], url_path='update')
    def update_doctor(self, request, pk=None):
        try:
            doctor = Doctor.objects.filter(id=pk).first()
            if not doctor:
                return self.failure_response(
                    data={},
                    message="doctor topilmadi",
                    status_code=404
                )
            serializer = self.get_serializer(data=request.data, instance=doctor)
            if serializer.is_valid():
                obj = serializer.save()
                return self.success_response(
                    data=serializers.DoctorListSerializer(obj).data,
                    message='malumot tahrirlandi'
                )
            return self.failure_response(
                data=serializer.errors,
                message='malumot tahrirlandi'
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )
    
    @swagger_auto_schema(
        tags=['Admin Doctors']
    )
    @action(detail=True, methods=['delete'], url_path='delete')
    def delete(self, request, pk=None):
        try:
            doctor = Doctor.objects.filter(id=pk).first()
            if not doctor:
                return self.failure_response(
                    data={},
                    message="doctor topilmadi",
                    status_code=404
                )
            doctor.delete()
            return self.success_response(
                data={},
                message='malumot ochirildi',
                status_code=204
            )
        except Exception as e:
            return self.error_response(
                data=str(e),
                message="xatolik"
            )
    
    

class DoctorExportView(views.APIView, ResponseMixin):
    permission_classes = [permissions.IsAuthenticated]
    
    @swagger_auto_schema(
        tags=['Admin Doctors']
    )
    def get(self, request):
        try:
            queryset = Doctor.objects.select_related(
                'district', 'place', 'user'
            ).all()
            
            queryset = self.filter_queryset(request, queryset)
            
            excel_file = self.create_excel_file(queryset)
            
            response = HttpResponse(
                excel_file,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="shifokorlar.xlsx"'
            response['Content-Length'] = len(excel_file)
            
            return response
            
        except Exception as e:
            return self.error_response(
                data=str(e),
            )
    
    def filter_queryset(self, request, queryset):
        district_id = request.query_params.get('district_id')
        if district_id:
            queryset = queryset.filter(district_id=district_id)
        
        place_id = request.query_params.get('place_id')
        if place_id:
            queryset = queryset.filter(place_id=place_id)
        
        sphere = request.query_params.get('sphere')
        if sphere:
            queryset = queryset.filter(sphere__icontains=sphere)
        
        work_place = request.query_params.get('work_place')
        if work_place:
            queryset = queryset.filter(work_place__icontains=work_place)
        
        user_id = request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        return queryset
    
    def create_excel_file(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Shifokorlar"
        
        header_fill = PatternFill(
            start_color="4472C4", 
            end_color="4472C4", 
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11, name='Arial')
        header_alignment = Alignment(
            horizontal="center", 
            vertical="center",
            wrap_text=True
        )
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            'ID',
            'Ism',
            'Familiya',
            'Telefon raqami',
            'Ish joyi',
            'Mutaxassisligi',
            "Qo'shimcha ma'lumot",
            'Tuman',
            'Joy',
            'Foydalanuvchi',
            'Longitude',
            'Latitude',
            "Qo'shimcha manzil"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        data_font = Font(size=10, name='Arial')
        data_alignment = Alignment(vertical="center", wrap_text=False)
        
        for row_num, doctor in enumerate(queryset, 2):
            cell = ws.cell(row=row_num, column=1)
            cell.value = doctor.id
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=2)
            cell.value = doctor.first_name
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=3)
            cell.value = doctor.last_name
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=4)
            cell.value = doctor.phone_number
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=5)
            cell.value = doctor.work_place
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=6)
            cell.value = doctor.sphere
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=7)
            cell.value = doctor.description if doctor.description else ''
            cell.font = data_font
            cell.alignment = alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=8)
            cell.value = doctor.district.name if doctor.district else ''
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=9)
            cell.value = doctor.place.name if doctor.place else ''
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=10)
            cell.value = doctor.user.username if doctor.user else ''
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=11)
            cell.value = doctor.longitude
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row_num, column=12)
            cell.value = doctor.latitude
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            cell = ws.cell(row=row_num, column=13)
            if doctor.extra_location:
                try:
                    cell.value = json.dumps(doctor.extra_location, ensure_ascii=False)
                except:
                    cell.value = str(doctor.extra_location)
            else:
                cell.value = ''
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 30
        
        ws.freeze_panes = 'A2'
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()